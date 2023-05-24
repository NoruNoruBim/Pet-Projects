from structures import *
from datetime import datetime
from nt import mkdir
from openpyxl.reader.excel import load_workbook
import zlib
import sys


#   class to work with data base
class DataBase:
    @show
    def __init__(self, db_name, layers_list):#  init layers, specified in layers_list from .*** file
        self.db_name = db_name
        
        db_file = open(self.db_name, 'rb')
        
        db_file.seek(0, 0)
        self.Header = Header(db_file)
        
        offsets = {decrypt(i.Enum_an_class) : decrypt(i.Offset_of_layer_header) for i in self.Header.header_offsets}

        self.NDB = Layer(db_file, abn_point, offsets[1], 'NDB_1') if 1 in layers_list and 1 in offsets else None
        
        self.VHF = Layer(db_file, abn_vhf, offsets[2], 'VHF_2') if 2 in layers_list and 2 in offsets else None
        
        '''
        github
        here is another 14 removed layers
        '''
        
        db_file.close()


    @show
    def write_to_file(self, filename, layers_list):#    just write specified layers back to .*** file
        copy = open(filename, 'wb')
        with open(self.db_name, 'rb') as orig:
            copy.write(orig.read())
        
        offsets = {decrypt(i.Enum_an_class) : decrypt(i.Offset_of_layer_header) for i in self.Header.header_offsets}
        offsets.update({0 : 0})#    Header offset
        
        layer_fields = vars(self)
        for layer in LAYER_TO_NUMBER:
            number = LAYER_TO_NUMBER[layer]
            if number in offsets and number in layers_list:
                layer_fields[layer].write_to_file(copy, offsets[number], layer+'_'+str(number))

        copy.close()

        with open(filename, 'rb+') as copy:#   update crc and file_size
            copy.seek(0, 0)
            new_length = len(copy.read())
            copy.seek(24, 0)
            copy.write(encrypt(new_length, 4))

            body_crc = self._countBodyCRC(copy)

            copy.seek(28, 0)
            copy.write(encrypt(int(float.fromhex(body_crc)), 4))

            head_crc = self._countHeadCRC(copy)

            copy.seek(36, 0)
            copy.write(encrypt(int(float.fromhex(head_crc)), 4))
        
        print('head_crc =', head_crc, 'body_crc =', body_crc, end='\n\n')
        

    @show
    def write_to_excel(self, layers_list):# write specified layers to excel files
        curTime = datetime.now().strftime("%Y%m%d_%H%M%S")

        FOLDER_NEW_NAME = self.db_name[:-4] + "_" + curTime + "/"
        try:
            mkdir(FOLDER_NEW_NAME)
            mkdir(FOLDER_NEW_NAME + 'index_arrs/')
        except:
            exit(" --- ERROR: Can't make folders")
        
        offsets = {decrypt(i.Enum_an_class) : decrypt(i.Offset_of_layer_header) for i in self.Header.header_offsets}
        
        layer_fields = vars(self)
        for layer in LAYER_TO_NUMBER:
            number = LAYER_TO_NUMBER[layer]
            if layer == 'Header':
                wb = load_workbook(TEMPLATES_FOLDER_NAME + TEMPLATES_FILES_NAMES[layer])
                sheet = wb.active
                
                layer_fields[layer].write_to_excel(sheet)
                
                wb.save(FOLDER_NEW_NAME + TEMPLATES_FILES_NAMES[layer])
            elif number in offsets and number in layers_list:
            
                wb_header = load_workbook(TEMPLATES_FOLDER_NAME + TEMPLATES_FILES_NAMES[layer + '_Header'])
                sheet_header = wb_header.active

                wb = load_workbook(TEMPLATES_FOLDER_NAME + TEMPLATES_FILES_NAMES[layer])
                sheet = wb.active
                
                prefix = str(LAYER_TO_NUMBER[layer]) if LAYER_TO_NUMBER[layer] >= 10 else '0' + str(LAYER_TO_NUMBER[layer])
                wb_index = load_workbook(TEMPLATES_FOLDER_NAME + 'index_arrs/' + prefix + '_index_arrs.xlsx')
                sheet_index = wb_index.active
                
                layer_fields[layer].write_to_excel(sheet_header, sheet, sheet_index, layer+'_'+str(number))
                
                wb_index.save(FOLDER_NEW_NAME + 'index_arrs/' + prefix + '_index_arrs.xlsx')
                wb.save(FOLDER_NEW_NAME + TEMPLATES_FILES_NAMES[layer])
                wb_header.save(FOLDER_NEW_NAME + TEMPLATES_FILES_NAMES[layer + '_Header'])

    
    def _countBodyCRC(self, db_file):
        db_file.seek(220)
        data = db_file.read()
        body_crc = hex(zlib.crc32(data) & 0xFFFFFFFF)
        return body_crc


    def _countHeadCRC(self, db_file):
        db_file.seek(0, 0)
        data = db_file.read(220)
        data1 = data[:36] + b'\x00\x00\x00\x00' + data[40:]
        head_crc = hex(zlib.crc32(data1) & 0xFFFFFFFF)
        return head_crc


    def _countCRC(self, db_file):#  count crc of .*** file like in requerments
        db_file.seek(220)
        data = db_file.read()
        body_crc = hex(zlib.crc32(data) & 0xFFFFFFFF)
        
        db_file.seek(0, 0)
        data = db_file.read(220)
        data1 = data[:36] + b'\x00\x00\x00\x00' + data[40:]
        head_crc = hex(zlib.crc32(data1) & 0xFFFFFFFF)

        return head_crc, body_crc


    #   shift all needed data after adding or removing objects from layer or index arrays
    def _do_shift(self, shift_index, shift_data, layer_cur, layer_fields, LAYER_TO_NUMBER, offsets):
        if shift_index == 0 and shift_data == 0:
            return
        
        '''
        github
        here is removed code
        '''

    @show
    def modify(self, excel_path, layers_list):#  read from excel data and paste into db object
        offsets = {decrypt(i.Enum_an_class) : decrypt(i.Offset_of_layer_header) for i in self.Header.header_offsets}
        
        layer_fields = vars(self)
        for layer in LAYER_TO_NUMBER:
            number = LAYER_TO_NUMBER[layer]
            if layer == 'Header':
                wb = load_workbook(excel_path + TEMPLATES_FILES_NAMES[layer])
                sheet = wb.active
                
                layer_fields[layer].modify(sheet)
                
            elif number in offsets and number in layers_list:
            
                wb_header = load_workbook(excel_path + TEMPLATES_FILES_NAMES[layer + '_Header'])
                sheet_header = wb_header.active

                wb = load_workbook(excel_path + TEMPLATES_FILES_NAMES[layer])
                sheet = wb.active
                
                prefix = str(LAYER_TO_NUMBER[layer]) if LAYER_TO_NUMBER[layer] >= 10 else '0' + str(LAYER_TO_NUMBER[layer])                
                try:
                    wb_index = load_workbook(excel_path + 'index_arrs/' + prefix + '_index_arrs.xlsx')
                    sheet_index = wb_index.active
                except:
                    sheet_index = None
                
                shift_index, shift_data = layer_fields[layer].modify(LAYER_TO_STRUCT[layer], sheet_header, sheet, sheet_index, layer+'_'+str(number))
                self._do_shift(shift_index, shift_data, layer, layer_fields, LAYER_TO_NUMBER, offsets)


#   user launch from console
if __name__ == '__main__':
    args = sys.argv
    assert not ('-extract' in args and '-build' in args), 'Use only one of "-extract" or "-build"'
    assert len(args) >= 3, 'Specify working mode: "-extract" or "-build" and path to DB'
    
    db_name = args[2]
    if '-extract' in args:
        layers_list = list(map(lambda x: int(x[1:]), args[3:])) if len(args) > 3 else [i for i in range(1, 17)]
        assert all(map(lambda x: type(x) == int and 0 <= x <= 16, layers_list)), 'Use correct layer numbers [0; 16]'
    
        db = DataBase(db_name, layers_list)#  Always read Header
        db.write_to_excel(layers_list)#  Always extract Header

    if '-build' in args:
        layers_list = list(map(lambda x: int(x[1:]), args[5:])) if len(args) > 5 else [i for i in range(1, 17)]
        assert all(map(lambda x: type(x) == int and 0 <= x <= 16, layers_list)), 'Use correct layer numbers [0; 16]'
    
        folder_name = args[3]
        copy_name = args[4]
        db = DataBase(db_name, [i for i in range(1, 17)])#    its possible to optimize this code (open only needed)
        db.modify(folder_name + '/', layers_list)
        db.write_to_file(copy_name, [i for i in range(0, 17)])
